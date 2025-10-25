#!/usr/bin/env python3
"""
Sistema de Análise de Dados - Indústria Têxtil
Conecta ao MySQL, analisa dados e gera relatório Excel formatado
"""
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Importar configurações
try:
    from config import MYSQL_CONFIG
except ImportError:
    print("❌ Erro: arquivo 'config.py' não encontrado!")
    print("   Copie 'config_example.py' para 'config.py' e configure.")
    sys.exit(1)


class AnaliseDadosTextil:
    """Classe principal para análise de dados da indústria têxtil"""
    
    def __init__(self):
        self.engine = None
        self.dados = {}
        self.excel_filename = None
        
    def conectar_mysql(self):
        """Conecta ao MySQL usando SQLAlchemy"""
        try:
            print("🔌 Conectando ao MySQL...")
            connection_string = (
                f"mysql+mysqlconnector://{MYSQL_CONFIG['user']}:{MYSQL_CONFIG['password']}"
                f"@{MYSQL_CONFIG['host']}:{MYSQL_CONFIG['port']}/{MYSQL_CONFIG['database']}"
                f"?charset={MYSQL_CONFIG.get('charset', 'utf8mb4')}"
            )
            self.engine = create_engine(connection_string)
            # Testar conexão
            with self.engine.connect():
                print("✅ Conectado ao MySQL com sucesso!")
                return True
        except SQLAlchemyError as e:
            print(f"❌ Erro ao conectar ao MySQL: {e}")
            return False
    
    def extrair_dados(self):
        """Extrai dados do MySQL"""
        print("\n📥 Extraindo dados do MySQL...")
        
        queries = {
            'vendas_por_produto': """
                SELECT 
                    p.categoria as 'Categoria',
                    p.tamanho as 'Tamanho',
                    p.cor as 'Cor',
                    COUNT(v.id) as 'Quantidade de Vendas',
                    SUM(v.quantidade) as 'Unidades Vendidas',
                    SUM(v.valor_total) as 'Valor Total (R$)'
                FROM vendas v
                JOIN produtos p ON v.produto_id = p.id
                GROUP BY p.categoria, p.tamanho, p.cor
            """,
            
            'producao_por_turno': """
                SELECT 
                    turno as 'Turno',
                    COUNT(*) as 'Quantidade de Produções',
                    SUM(quantidade_produzida) as 'Total Produzido (unidades)',
                    AVG(tempo_producao_horas) as 'Tempo Médio (horas)',
                    qualidade as 'Qualidade',
                    COUNT(*) as 'Número de Registros'
                FROM producao
                GROUP BY turno, qualidade
            """,
            
            'tecidos_mais_usados': """
                SELECT 
                    t.tipo as 'Tipo de Tecido',
                    t.cor as 'Cor',
                    COUNT(cm.id) as 'Vezes Usado',
                    SUM(cm.quantidade_usada) as 'Total Usado (metros)'
                FROM consumo_materiais cm
                JOIN tecidos t ON cm.tecido_id = t.id
                WHERE cm.tecido_id IS NOT NULL
                GROUP BY t.tipo, t.cor
                ORDER BY SUM(cm.quantidade_usada) DESC
            """,
            
            'agulhas_mais_usadas': """
                SELECT 
                    a.tipo as 'Tipo de Agulha',
                    a.tamanho as 'Tamanho',
                    COUNT(cm.id) as 'Vezes Usado',
                    SUM(cm.quantidade_usada) as 'Total de Agulhas (unidades)'
                FROM consumo_materiais cm
                JOIN agulhas a ON cm.agulha_id = a.id
                WHERE cm.agulha_id IS NOT NULL
                GROUP BY a.tipo, a.tamanho
                ORDER BY COUNT(cm.id) DESC
            """,
            
            'linhas_mais_usadas': """
                SELECT 
                    rl.tipo as 'Tipo de Linha',
                    rl.cor as 'Cor',
                    COUNT(cm.id) as 'Vezes Usado',
                    SUM(cm.quantidade_usada) as 'Total Usado (metros)'
                FROM consumo_materiais cm
                JOIN rolos_linha rl ON cm.rolo_linha_id = rl.id
                WHERE cm.rolo_linha_id IS NOT NULL
                GROUP BY rl.tipo, rl.cor
                ORDER BY SUM(cm.quantidade_usada) DESC
            """,
            
            'manutencao_por_tipo': """
                SELECT 
                    tipo_manutencao as 'Tipo de Manutenção',
                    COUNT(*) as 'Quantidade',
                    SUM(custo) as 'Custo Total (R$)',
                    AVG(custo) as 'Custo Médio (R$)',
                    SUM(tempo_parada_horas) as 'Tempo Parada Total (horas)',
                    AVG(tempo_parada_horas) as 'Tempo Parada Médio (horas)'
                FROM manutencao_maquinas
                GROUP BY tipo_manutencao
            """,
            
            'producao_por_setor': """
                SELECT 
                    f.setor as 'Setor',
                    COUNT(DISTINCT f.id) as 'Número de Funcionários',
                    COUNT(p.id) as 'Quantidade de Produções',
                    SUM(p.quantidade_produzida) as 'Total Produzido (unidades)'
                FROM funcionarios f
                LEFT JOIN producao p ON p.operador = f.nome
                GROUP BY f.setor
            """,
            
            'vendas_por_forma_pagamento': """
                SELECT 
                    forma_pagamento as 'Forma de Pagamento',
                    COUNT(*) as 'Quantidade de Vendas',
                    SUM(valor_total) as 'Valor Total (R$)',
                    AVG(valor_total) as 'Ticket Médio (R$)'
                FROM vendas
                GROUP BY forma_pagamento
            """,
            
            'estoque_atual': """
                SELECT 
                    'Rolos de Linha' as 'Item',
                    COUNT(*) as 'Quantidade de Itens',
                    SUM(quantidade_estoque) as 'Estoque Total (unidades)'
                FROM rolos_linha
                UNION ALL
                SELECT 
                    'Agulhas' as 'Item',
                    COUNT(*) as 'Quantidade de Itens',
                    SUM(quantidade_estoque) as 'Estoque Total (unidades)'
                FROM agulhas
                UNION ALL
                SELECT 
                    'Tecidos' as 'Item',
                    COUNT(*) as 'Quantidade de Itens',
                    SUM(metragem_estoque) as 'Estoque Total (metros)'
                FROM tecidos
            """,
            
            'top_clientes': """
                SELECT 
                    c.nome as 'Nome do Cliente',
                    c.cidade as 'Cidade',
                    c.estado as 'Estado',
                    COUNT(v.id) as 'Número de Compras',
                    SUM(v.valor_total) as 'Valor Total Comprado (R$)'
                FROM clientes c
                JOIN vendas v ON c.id = v.cliente_id
                GROUP BY c.id, c.nome, c.cidade, c.estado
                ORDER BY SUM(v.valor_total) DESC
                LIMIT 20
            """
        }
        
        for nome, query in queries.items():
            try:
                print(f"  → {nome}...")
                self.dados[nome] = pd.read_sql(query, self.engine)
                

                if not self.dados[nome].empty:
                    for col in self.dados[nome].columns:
                        if self.dados[nome][col].dtype == 'object': 
                            self.dados[nome][col] = self.dados[nome][col].apply(
                                lambda x: x.encode('latin1').decode('utf8') if isinstance(x, str) and 'Ã' in x else x
                            )
            except Exception as e:
                print(f"    ⚠️  Erro: {e}")
                self.dados[nome] = pd.DataFrame()
        
        print(f"✅ {len(self.dados)} conjuntos de dados extraídos!")
    
    def gerar_relatorio_local(self):
        """Gera relatório Excel local completo com formatação"""
        print("\n💾 Gerando relatório Excel formatado...")
        
        self.excel_filename = f'relatorio_textil_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Criar Excel com pandas
        with pd.ExcelWriter(self.excel_filename, engine='openpyxl') as writer:
            # Dashboard primeiro
            self._criar_dashboard_excel(writer)
            
            # Exportar cada dataset
            for nome, df in self.dados.items():
                if not df.empty:
                    sheet_name = nome.replace('_', ' ').title()[:31]

                    titulo_df = pd.DataFrame([[f'📊 {sheet_name}']], columns=[''])
                    titulo_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        
        # Aplicar formatação
        self._formatar_excel()
        
        print(f"✅ Relatório Excel salvo: {self.excel_filename}")
        return self.excel_filename
    
    def _criar_dashboard_excel(self, writer):
        """Cria aba de dashboard no Excel com tops e gráficos"""
        try:
            # Buscar dados para resumo geral
            totais = {
                'vendas': len(pd.read_sql("SELECT id FROM vendas", self.engine)),
                'producoes': len(pd.read_sql("SELECT id FROM producao", self.engine)),
                'clientes': len(pd.read_sql("SELECT id FROM clientes", self.engine)),
                'funcionarios': len(pd.read_sql("SELECT id FROM funcionarios", self.engine)),
                'fornecedores': len(pd.read_sql("SELECT id FROM fornecedores", self.engine)),
                'valor_vendas': pd.read_sql("SELECT SUM(valor_total) as total FROM vendas", self.engine)["total"][0],
                'custo_manutencao': pd.read_sql("SELECT SUM(custo) as total FROM manutencao_maquinas", self.engine)["total"][0],
            }
            

            top_clientes = pd.read_sql("""
                SELECT 
                    c.nome,
                    SUM(v.valor_total) as valor_total
                FROM clientes c
                JOIN vendas v ON c.id = v.cliente_id
                GROUP BY c.id, c.nome
                ORDER BY valor_total DESC
                LIMIT 3
            """, self.engine)
            
            top_pagamentos = pd.read_sql("""
                SELECT 
                    forma_pagamento,
                    COUNT(*) as quantidade,
                    SUM(valor_total) as valor_total
                FROM vendas
                GROUP BY forma_pagamento
                ORDER BY valor_total DESC
                LIMIT 3
            """, self.engine)
            
            top_turnos = pd.read_sql("""
                SELECT 
                    turno,
                    COUNT(*) as quantidade,
                    SUM(quantidade_produzida) as total_produzido
                FROM producao
                GROUP BY turno
                ORDER BY total_produzido DESC
                LIMIT 3
            """, self.engine)
            
            if not top_turnos.empty and 'turno' in top_turnos.columns:
                top_turnos['turno'] = top_turnos['turno'].apply(
                    lambda x: x.encode('latin1').decode('utf8') if isinstance(x, str) and 'Ã' in x else x
                )
            
            manutencao_cara = pd.read_sql("""
                SELECT 
                    tipo_manutencao,
                    SUM(custo) as custo_total,
                    SUM(tempo_parada_horas) as tempo_total_horas
                FROM manutencao_maquinas
                GROUP BY tipo_manutencao
                ORDER BY custo_total DESC
                LIMIT 1
            """, self.engine)
            
            dashboard_data = {
                'Indicador': [
                    f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}',
                    '',
                    '📈 RESUMO GERAL',
                    '',
                    'Total de Vendas',
                    'Total de Produções',
                    'Total de Clientes',
                    'Total de Funcionários',
                    'Total de Fornecedores',
                    '',
                    '💰 FINANCEIRO',
                    '',
                    'Valor Total Vendas',
                    'Custo Total Manutenção',
                ],
                'Valor': [
                    '',
                    '',
                    '',
                    '',
                    f"{totais['vendas']:,} vendas",
                    f"{totais['producoes']:,} produções",
                    f"{totais['clientes']:,} clientes",
                    f"{totais['funcionarios']:,} funcionários",
                    f"{totais['fornecedores']:,} fornecedores",
                    '',
                    '',
                    '',
                    f"R$ {totais['valor_vendas']:,.2f}",
                    f"R$ {totais['custo_manutencao']:,.2f}",
                ]
            }
            
            df_dashboard = pd.DataFrame(dashboard_data)
            df_dashboard.to_excel(writer, sheet_name='📊 Dashboard', index=False, startrow=0)
            
            top_clientes_data = {
                'Indicador': ['', '🏆 TOP 3 CLIENTES', ''] + 
                             [f"{i+1}º - {row['nome']}" for i, row in top_clientes.iterrows()],
                'Valor': ['', '', ''] + 
                         [f"R$ {row['valor_total']:,.2f}" for _, row in top_clientes.iterrows()]
            }
            df_top_clientes = pd.DataFrame(top_clientes_data)
            df_top_clientes.to_excel(writer, sheet_name='📊 Dashboard', index=False, startrow=0, startcol=4)
            
            top_pag_data = {
                'Indicador': ['', '💳 TOP 3 FORMAS PAGAMENTO', ''] + 
                             [f"{i+1}º - {row['forma_pagamento']}" for i, row in top_pagamentos.iterrows()],
                'Nº de Vendas': ['', '', ''] + 
                         [f"{int(row['quantidade']):,} vendas" for _, row in top_pagamentos.iterrows()],
                'Valor Total (R$)': ['', '', ''] + 
                         [f"R$ {row['valor_total']:,.2f}" for _, row in top_pagamentos.iterrows()]
            }
            df_top_pag = pd.DataFrame(top_pag_data)
            df_top_pag.to_excel(writer, sheet_name='📊 Dashboard', index=False, startrow=0, startcol=7)
            
            manutencao_data = {
                'Indicador': [
                    '🔧 MANUTENÇÃO MAIS CARA',
                    '',
                    'Tipo',
                    'Custo Total',
                    'Tempo de Parada Total',
                ],
                'Valor': [
                    '',
                    '',
                    manutencao_cara['tipo_manutencao'].iloc[0] if not manutencao_cara.empty else 'N/A',
                    f"R$ {manutencao_cara['custo_total'].iloc[0]:,.2f}" if not manutencao_cara.empty else 'N/A',
                    f"{manutencao_cara['tempo_total_horas'].iloc[0]:,.0f} horas" if not manutencao_cara.empty else 'N/A',
                ]
            }
            df_manutencao = pd.DataFrame(manutencao_data)
            df_manutencao.to_excel(writer, sheet_name='📊 Dashboard', index=False, startrow=9, startcol=4)
            
            top_turnos_data = {
                'Indicador': ['', '🕐 TOP 3 TURNOS', ''] + 
                             [f"{i+1}º - {row['turno']}" for i, row in top_turnos.iterrows()],
                'Total Produzido (unid.)': ['', '', ''] + 
                         [f"{int(row['total_produzido']):,} unidades" for _, row in top_turnos.iterrows()]
            }
            df_top_turnos = pd.DataFrame(top_turnos_data)
            df_top_turnos.to_excel(writer, sheet_name='📊 Dashboard', index=False, startrow=0, startcol=11)
            

            top5_clientes = pd.read_sql("""
                SELECT 
                    c.nome,
                    SUM(v.valor_total) as valor_total
                FROM clientes c
                JOIN vendas v ON c.id = v.cliente_id
                GROUP BY c.id, c.nome
                ORDER BY valor_total DESC
                LIMIT 5
            """, self.engine)
            
            top5_pagamentos = pd.read_sql("""
                SELECT 
                    forma_pagamento,
                    SUM(valor_total) as valor_total
                FROM vendas
                GROUP BY forma_pagamento
                ORDER BY valor_total DESC
                LIMIT 5
            """, self.engine)
            
            top3_turnos = pd.read_sql("""
                SELECT 
                    turno,
                    SUM(quantidade_produzida) as total_produzido
                FROM producao
                GROUP BY turno
                ORDER BY total_produzido DESC
                LIMIT 3
            """, self.engine)
            
            if not top3_turnos.empty and 'turno' in top3_turnos.columns:
                top3_turnos['turno'] = top3_turnos['turno'].apply(
                    lambda x: x.encode('latin1').decode('utf8') if isinstance(x, str) and 'Ã' in x else x
                )
            
            start_row = 50
            

            df_grafico_clientes = pd.DataFrame({
                'Cliente': top5_clientes['nome'],
                'Valor Total (R$)': top5_clientes['valor_total']
            })
            df_grafico_clientes.to_excel(writer, sheet_name='📊 Dashboard', index=False, 
                                        startrow=start_row, startcol=0)
            
            df_grafico_pag = pd.DataFrame({
                'Forma de Pagamento': top5_pagamentos['forma_pagamento'],
                'Valor Total (R$)': top5_pagamentos['valor_total']
            })
            df_grafico_pag.to_excel(writer, sheet_name='📊 Dashboard', index=False, 
                                   startrow=start_row, startcol=4)
            
            df_grafico_turnos = pd.DataFrame({
                'Turno': top3_turnos['turno'],
                'Total Produzido (unidades)': top3_turnos['total_produzido']
            })
            df_grafico_turnos.to_excel(writer, sheet_name='📊 Dashboard', index=False, 
                                      startrow=start_row, startcol=8)
            
        except Exception as e:
            print(f"    ⚠️  Erro ao criar dashboard: {e}")
    
    def _formatar_excel(self):
        """Aplica formatação ao Excel"""
        print("  → Aplicando formatação...")
        
        wb = load_workbook(self.excel_filename)
        
        cor_header = PatternFill(start_color="00B2A4", end_color="00B2A4", fill_type="solid")
        cor_titulo = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        cor_secao = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        font_header = Font(bold=True, size=12, color="FFFFFF")
        font_secao = Font(bold=True, size=11)
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if sheet_name == '📊 Dashboard':
                self._formatar_dashboard(ws, font_secao, align_left, cor_secao)
            else:
                self._formatar_aba_dados(ws, font_header, cor_titulo, cor_header, 
                                        align_center, border_thin)
        
        wb.save(self.excel_filename)
        print("  → Formatação aplicada!")
    
    def _formatar_dashboard(self, ws, font_secao, align_left, cor_secao):
        """Formata aba do dashboard e adiciona gráficos"""
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['L'].width = 25
        ws.column_dimensions['M'].width = 25
        
        secoes_linha = [3, 11]
        for row in secoes_linha:
            cell = ws.cell(row=row, column=1)
            cell.font = font_secao
            cell.fill = cor_secao
            cell.alignment = align_left
        
        cell = ws.cell(row=10, column=5)
        cell.font = font_secao
        cell.fill = cor_secao
        cell.alignment = align_left
        
        tops_cols = [5, 8]
        for col in tops_cols:
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="00B2A4", end_color="00B2A4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(1, ws.max_row + 1):
            for col in [1, 5, 8]:
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).alignment = align_left
        
        for row in range(52, 57):
            cell = ws.cell(row=row, column=2)
            cell.number_format = 'R$ #,##0.00'
        
        for row in range(52, 57):
            cell = ws.cell(row=row, column=6)
            cell.number_format = 'R$ #,##0.00'
        
        for row in range(52, 55):
            cell = ws.cell(row=row, column=10)
            cell.number_format = '#,##0'
        
        self._adicionar_graficos_pizza(ws)
    
    def _adicionar_graficos_pizza(self, ws):
        """Adiciona gráficos de pizza no dashboard"""
        try:
            chart1 = PieChart()
            chart1.title = "🏆 Top 5 Clientes"
            chart1.style = 10
            chart1.height = 10
            chart1.width = 15
            
            data1 = Reference(ws, min_col=2, min_row=51, max_row=56)
            labels1 = Reference(ws, min_col=1, min_row=52, max_row=56)
            
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(labels1)
            chart1.legend = None
            
            chart1.dataLabels = DataLabelList()
            chart1.dataLabels.showPercent = True
            chart1.dataLabels.showVal = False
            chart1.dataLabels.showCatName = False
            
            ws.add_chart(chart1, "A18")
            
            chart2 = PieChart()
            chart2.title = "💳 Top 5 Formas de Pagamento"
            chart2.style = 11
            chart2.height = 10
            chart2.width = 15
            
            data2 = Reference(ws, min_col=6, min_row=51, max_row=56)
            labels2 = Reference(ws, min_col=5, min_row=52, max_row=56)
            
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(labels2)
            chart2.legend = None
            
            chart2.dataLabels = DataLabelList()
            chart2.dataLabels.showPercent = True
            chart2.dataLabels.showVal = False
            chart2.dataLabels.showCatName = False
            
            ws.add_chart(chart2, "E18")
            
            chart3 = PieChart()
            chart3.title = "🕐 Top 3 Turnos de Produção"
            chart3.style = 12
            chart3.height = 10
            chart3.width = 15
            
            data3 = Reference(ws, min_col=10, min_row=51, max_row=54)
            labels3 = Reference(ws, min_col=9, min_row=52, max_row=54)
            
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(labels3)
            chart3.legend = None
            
            chart3.dataLabels = DataLabelList()
            chart3.dataLabels.showPercent = True
            chart3.dataLabels.showVal = False
            chart3.dataLabels.showCatName = False
            
            ws.add_chart(chart3, "I18")
            
            print("  → Gráficos de pizza adicionados!")
            
        except Exception as e:
            print(f"    ⚠️  Erro ao adicionar gráficos: {e}")
    
    def _formatar_aba_dados(self, ws, font_header, cor_titulo, cor_header,
                            align_center, border_thin):
        """Formata aba de dados"""
        ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = cor_titulo
        title_cell.alignment = align_center
        
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = font_header
            cell.fill = cor_header
            cell.alignment = align_center
            cell.border = border_thin
        
        ws.auto_filter.ref = f'A3:{get_column_letter(ws.max_column)}{ws.max_row}'
        
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, min(ws.max_row + 1, 100)):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'A4'
    
    def executar(self):
        """Executa análise completa"""
        print("="*70)
        print("🏭 SISTEMA DE ANÁLISE - INDÚSTRIA TÊXTIL")
        print("="*70)
        
        if not self.conectar_mysql():
            return False
        
        self.extrair_dados()
        excel_file = self.gerar_relatorio_local()
        
        if self.engine:
            self.engine.dispose()
            print("\n✅ Análise concluída com sucesso!")
            print(f"   📊 Excel gerado: {excel_file}")
            print("\n💡 Dica: Para enviar ao Google Sheets, faça upload manual em:")
            print("      https://drive.google.com/")
        
        return True


if __name__ == '__main__':
    analise = AnaliseDadosTextil()
    success = analise.executar()
    sys.exit(0 if success else 1)

